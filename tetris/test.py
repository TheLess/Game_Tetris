from openpyxl import load_workbook
wb = load_workbook(r'e:\work\Code\Tetris\Block.xlsx', data_only=True, read_only=True)
ws = wb.active
rows = list(ws.iter_rows(values_only=True))
print(rows[0])  # 表头
for row in rows[1:6]:
    print(row)