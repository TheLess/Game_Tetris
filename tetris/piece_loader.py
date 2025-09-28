from __future__ import annotations

import logging
from dataclasses import dataclass
from pathlib import Path
from typing import Iterable, List, Sequence

from openpyxl import load_workbook

from .piece import Piece, build_matrix_from_rows


logger = logging.getLogger(__name__)


class PieceLoadError(Exception):
    """加载方块配置失败。"""


@dataclass(slots=True)
class RawRow:
    shape_id: str
    display_name: str
    cells: int | None
    allow_rotate: bool
    allow_mirror: bool
    spawn_weight: float
    color_hex: str | None
    matrix_size: int
    rows: Sequence[str]
    notes: str | None

    def to_piece(self) -> Piece:
        matrix = build_matrix_from_rows(self.rows, self.matrix_size)
        piece = Piece(
            shape_id=self.shape_id,
            display_name=self.display_name or self.shape_id,
            matrix=matrix,
            allow_rotate=self.allow_rotate,
            allow_mirror=self.allow_mirror,
            spawn_weight=self.spawn_weight,
            color_hex=self.color_hex,
            notes=self.notes,
        )
        if self.cells is not None and piece.cell_count != self.cells:
            raise PieceLoadError(
                f"方块 {self.shape_id} 的 Cells={self.cells} 与矩阵实际格子数 {piece.cell_count} 不一致"
            )
        return piece


def load_pieces_from_excel(path: str | Path, *, sheet_name: str | None = None) -> List[Piece]:
    """从 Excel 配置表加载所有方块。"""

    file_path = Path(path)
    if not file_path.exists():
        logger.error("未找到配置文件: %s", file_path)
        raise PieceLoadError(f"未找到配置文件: {file_path}")

    logger.debug("开始从 Excel 加载方块: path=%s sheet=%s", file_path, sheet_name or "<active>")
    try:
        workbook = load_workbook(filename=file_path, data_only=True, read_only=True)
    except Exception as exc:  # pragma: no cover - openpyxl 内部异常
        logger.exception("读取 Excel 文件失败: path=%s", file_path)
        raise PieceLoadError(f"读取 Excel 文件失败: {exc}") from exc

    try:
        worksheet = workbook[sheet_name] if sheet_name else workbook.active
    except KeyError as exc:
        logger.error("未找到工作表: %s", sheet_name)
        raise PieceLoadError(f"未找到工作表: {sheet_name}") from exc

    rows_iterator = worksheet.iter_rows(values_only=True)
    try:
        header = next(rows_iterator)
    except StopIteration as exc:
        raise PieceLoadError("Excel 表为空") from exc

    header_map = _build_header_map(header)
    required_headers = [
        "ShapeID",
        "DisplayName",
        "Cells",
        "AllowRotate",
        "AllowMirror",
        "SpawnWeight",
        "ColorHex",
        "MatrixSize",
        "Row1",
        "Row2",
        "Row3",
        "Row4",
        "Row5",
        "Notes",
    ]
    missing = [name for name in required_headers if name not in header_map]
    if missing:
        logger.error("缺少必要列: %s", ", ".join(missing))
        raise PieceLoadError(f"缺少必要列: {', '.join(missing)}")

    pieces: List[Piece] = []
    for row_index, raw_values in enumerate(rows_iterator, start=2):
        try:
            raw_row = _parse_raw_row(raw_values, header_map)
        except PieceLoadError as exc:
            logger.error("第 %s 行解析失败: %s", row_index, exc)
            raise PieceLoadError(f"第 {row_index} 行解析失败: {exc}") from exc
        if raw_row is None:
            logger.debug("第 %s 行未提供有效 ShapeID，跳过。", row_index)
            continue
        try:
            pieces.append(raw_row.to_piece())
        except PieceLoadError as exc:
            logger.error("第 %s 行数据非法: %s", row_index, exc)
            raise PieceLoadError(f"第 {row_index} 行数据非法: {exc}") from exc
        else:
            logger.debug(
                "读取方块: row=%s shape_id=%s allow_rotate=%s allow_mirror=%s",  # noqa: TRY401
                row_index,
                raw_row.shape_id,
                raw_row.allow_rotate,
                raw_row.allow_mirror,
            )

    if not pieces:
        logger.warning("未在配置中读取到任何方块: path=%s", file_path)
        raise PieceLoadError("未读取到任何方块数据")

    logger.info("成功加载 %s 个方块: path=%s", len(pieces), file_path)
    return pieces


def _build_header_map(header_row: Sequence[str | None]) -> dict[str, int]:
    mapping: dict[str, int] = {}
    for idx, name in enumerate(header_row):
        if name is None:
            continue
        normalized = str(name).strip()
        if normalized:
            mapping[normalized] = idx
    return mapping


def _parse_raw_row(values: Sequence[object], header_map: dict[str, int]) -> RawRow | None:
    shape_id = _get_cell(values, header_map, "ShapeID")
    if not shape_id:
        return None
    shape_id = str(shape_id).strip()
    if not shape_id:
        return None

    display_name = str(_get_cell(values, header_map, "DisplayName") or shape_id).strip()
    cells_value = _get_cell(values, header_map, "Cells")
    cells = int(cells_value) if cells_value not in (None, "") else None

    allow_rotate = _parse_bool(_get_cell(values, header_map, "AllowRotate"), default=True)
    allow_mirror = _parse_bool(_get_cell(values, header_map, "AllowMirror"), default=True)

    spawn_weight_value = _get_cell(values, header_map, "SpawnWeight")
    spawn_weight = float(spawn_weight_value) if spawn_weight_value not in (None, "") else 1.0
    if spawn_weight <= 0:
        raise PieceLoadError(f"方块 {shape_id} 的 SpawnWeight 必须大于 0")

    color_hex_value = _get_cell(values, header_map, "ColorHex")
    color_hex = str(color_hex_value).strip() if color_hex_value else None
    if color_hex == "":
        color_hex = None

    matrix_size_value = _get_cell(values, header_map, "MatrixSize")
    if matrix_size_value in (None, ""):
        raise PieceLoadError(f"方块 {shape_id} 的 MatrixSize 不能为空")
    matrix_size = int(matrix_size_value)

    row_fields = ["Row1", "Row2", "Row3", "Row4", "Row5"]
    rows: List[str] = []
    for field in row_fields:
        value = _get_cell(values, header_map, field)
        rows.append(str(value or "0" * matrix_size))

    notes_value = _get_cell(values, header_map, "Notes")
    notes = str(notes_value).strip() if notes_value else None

    return RawRow(
        shape_id=shape_id,
        display_name=display_name,
        cells=cells,
        allow_rotate=allow_rotate,
        allow_mirror=allow_mirror,
        spawn_weight=spawn_weight,
        color_hex=color_hex,
        matrix_size=matrix_size,
        rows=rows,
        notes=notes,
    )


def _get_cell(values: Sequence[object], header_map: dict[str, int], column: str) -> object:
    idx = header_map[column]
    return values[idx] if idx < len(values) else None


def _parse_bool(value: object, *, default: bool) -> bool:
    if value in (None, ""):
        return default
    if isinstance(value, bool):
        return value
    text = str(value).strip().lower()
    if text in {"true", "1", "yes", "y"}:
        return True
    if text in {"false", "0", "no", "n"}:
        return False
    raise PieceLoadError(f"无法解析布尔值: {value}")
